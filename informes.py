from time import strftime
from typing import Any
import pandas as pd
import openpyxl

from conexion_sqlite import Comunicacion

class Informes():
    def __init__(self):
        self.bd = Comunicacion()
    def save_libros(self):
        datos = self.bd.show_laminas()
        i = -1
        remitente,cantidad,niveleducativo,condicionlibro = [],[],[],[]
        autor,editorial,añoedicion,titulo,añorecepcion= [],[],[],[],[]
        for dato in datos:
            remitente.append(dato[0])
            añorecepcion.append(dato[1])
            niveleducativo.append(dato[2])
            titulo.append(dato[3])
            autor.append(dato[4])
            editorial.append(dato[5])
            añoedicion.append(dato[6])
            condicionlibro.append(dato[7])
            cantidad.append(dato[8])
        fecha = str(strftime('%d-%m-%y_%H-%M-%S'))
        df_datos = {'Remitente': remitente, 'Año de Entrega': añorecepcion,
                    'Nivel Educativo': niveleducativo, 'Titulo': titulo,
                    'Autor': autor, 'Editorial': editorial, 
                    'Año de Edicion': añoedicion, 
                    'Condicion': condicionlibro, 'Cantidad': cantidad}
        df = pd.DataFrame(df_datos)
        """ 
        index=None ,columns=['Remitente',
        'Año de Entrega', 'Nivel Educativo',
        'Titulo', 'Autor', 'Editorial',
        'Año de Edicion', 'Condicion', 'Cantidad'])
        """
        name_xlsx = f'LIBROS {fecha}.xlsx'
        df.to_excel(name_xlsx, sheet_name='libros')
        workbook = openpyxl.load_workbook(name_xlsx)
        sheet = workbook['libros']
        sheet.delete_cols(1)
        workbook.save(name_xlsx)
        archivo_excel = pd.read_excel(name_xlsx)
        # print(archivo_excel[['Cantidad', 'Nivel Educativo', 'Condicion']])
        tabla_pivote = archivo_excel.pivot_table(index='Nivel Educativo',
                                                 columns='Condicion', 
                                                 values='Cantidad', aggfunc='sum')
        tabla_pivote.to_excel('libros_estado.xlsx', startrow=2, startcol=2, sheet_name='report')
        
    def save_laminas(self):
        l_datos = self.bd.show_laminas()
        i = -1
        remitente,cantidad,niveleducativo,condicionlamina = [],[],[],[]
        codigo,titulo,añorecepcion= [],[],[]
        for dato in l_datos:
            codigo.append(dato[0])
            remitente.append(dato[1])
            añorecepcion.append(dato[2])
            niveleducativo.append(dato[3])
            titulo.append(dato[4])
            condicionlamina.append(dato[5])
            cantidad.append(dato[6])
        fecha = str(strftime('%d-%m-%y_%H-%M-%S'))
        df_datos = {'Codigo':codigo, 'Remitente': remitente, 
                    'Año de Entrega': añorecepcion, 'Nivel Educativo': niveleducativo, 
                    'Titulo': titulo, 'Condicion': condicionlamina, 'Cantidad': cantidad}
        df = pd.DataFrame(df_datos)
        name_xlsx = f'LAMINAS {fecha}.xlsx'
        df.to_excel(name_xlsx, sheet_name='laminas')
        workbook = openpyxl.load_workbook(name_xlsx)
        sheet = workbook['laminas']
        sheet.delete_cols(1)
        workbook.save(name_xlsx)
        """
        archivo_excel = pd.read_excel(name_xlsx)
        # print(archivo_excel[['Cantidad', 'Nivel Educativo', 'Condicion']])
        tabla_pivote = archivo_excel.pivot_table(index='Nivel Educativo',
                                                 columns='Condicion', 
                                                 values='Cantidad', aggfunc='sum')
        tabla_pivote.to_excel('libros_estado.xlsx', startrow=2, startcol=2, sheet_name='report')
        """ 
        
    def save_nomina(self):
        l_alumnos = self.bd.show_alumnos()
        i = -1
        alumnoid,alumno,sexo = [],[],[]
        nivel,grado,seccion = [],[],[]
        for datos_alumno in l_alumnos:
            alumnoid.append(datos_alumno[0])
            alumno.append(datos_alumno[1])
            sexo.append(datos_alumno[2])
            nivel.append(datos_alumno[3])
            grado.append(datos_alumno[4])
            seccion.append(datos_alumno[5])
        fecha = str(strftime('%d-%m-%y_%H-%M-%S'))
        dicionario_alumno = {'AlumnoId': alumnoid, 'Alumno': alumno,
                    'Sexo': sexo, 'Nivel': nivel,
                    'Grado': grado, 'Seccion': seccion}
        data_alumnos = pd.DataFrame(dicionario_alumno)
        """ 
        index=None ,columns=['Remitente',
        'Año de Entrega', 'Nivel Educativo',
        'Titulo', 'Autor', 'Editorial',
        'Año de Edicion', 'Condicion', 'Cantidad'])
        """
        name_xlsx = f'LIBROS {fecha}.xlsx'
        data_alumnos.to_excel(name_xlsx, sheet_name='libros')
        workbook = openpyxl.load_workbook(name_xlsx)
        sheet = workbook['libros']
        sheet.delete_cols(1)
        workbook.save(name_xlsx)
        archivo_excel = pd.read_excel(name_xlsx)
        # print(archivo_excel[['Cantidad', 'Nivel Educativo', 'Condicion']])
        tabla_pivote = archivo_excel.pivot_table(index='Nivel Educativo',columns='Condicion')
        