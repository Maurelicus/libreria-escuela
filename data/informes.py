from time import strftime
from typing import Any
import pandas as pd
import openpyxl

from data.conexion_sqlite import Comunicacion

class Informes():
    def __init__(self):
        self.bd = Comunicacion()
        
    def save_libros(self):
        datos = self.bd.show_libros()
        i = -1
        titulo,autor,editorial,añoedicion,categoria,cantidad = [],[],[],[],[],[]
        remitente,niveleducativo,condicionlibro,añorecepcion= [],[],[],[]
        for dato in datos:
            titulo.append(dato[0])
            autor.append(dato[1])
            editorial.append(dato[2])
            añoedicion.append(dato[3])
            categoria.append(dato[4])
            cantidad.append(dato[5])
            remitente.append(dato[6])
            niveleducativo.append(dato[7])
            condicionlibro.append(dato[8])
            añorecepcion.append(dato[9])
        df_datos = {
            'Titulo': titulo,'Autor': autor,'Editorial': editorial,
            'Año de Edicion': añoedicion,'Categoria': categoria,'Cantidad': cantidad,
            'Remitente': remitente,'Nivel Educativo': niveleducativo, 
            'Condicion': condicionlibro,'Año de Recepcion': añorecepcion}
        
        fecha = str(strftime('%d-%m-%y_%H-%M-%S'))
        name_xlsx = f'LIBROS {fecha}.xlsx'
        df = pd.DataFrame(df_datos)
        df.to_excel(name_xlsx, sheet_name='libros')
        workbook = openpyxl.load_workbook(name_xlsx)
        sheet = workbook['libros']
        sheet.delete_cols(1)
        workbook.save(name_xlsx)
        
    def save_bajalibros(self):
        datos = self.bd.search_libros('Tipo',24)
        i = -1
        titulo,autor,editorial,añoedicion,categoria,cantidad = [],[],[],[],[],[]
        remitente,niveleducativo,condicionlibro,añorecepcion= [],[],[],[]
        for dato in datos:
            titulo.append(dato[0])
            autor.append(dato[1])
            editorial.append(dato[2])
            añoedicion.append(dato[3])
            categoria.append(dato[4])
            cantidad.append(dato[5])
            remitente.append(dato[6])
            niveleducativo.append(dato[7])
            condicionlibro.append(dato[8])
            añorecepcion.append(dato[9])
        df_datos = {
            'Titulo': titulo,'Autor': autor,'Editorial': editorial,
            'Año de Edicion': añoedicion,'Categoria': categoria,'Cantidad': cantidad,
            'Remitente': remitente,'Nivel Educativo': niveleducativo, 
            'Condicion': condicionlibro,'Año de Recepcion': añorecepcion}
        
        fecha = str(strftime('%d-%m-%y_%H-%M-%S'))
        name_xlsx = f'LIBROS_BAJA {fecha}.xlsx'
        df = pd.DataFrame(df_datos)
        df.to_excel(name_xlsx, sheet_name='libros_baja')
        workbook = openpyxl.load_workbook(name_xlsx)
        sheet = workbook['libros_baja']
        sheet.delete_cols(1)
        workbook.save(name_xlsx)
        
    def save_laminas(self):
        l_datos = self.bd.show_laminas()
        i = -1
        titulo,categoria,codigo,cantidad = [],[],[],[]
        remitente,condicion,niveleducativo,añorecepcion= [],[],[],[]
        for dato in l_datos:
            titulo.append(dato[0])
            categoria.append(dato[1])
            codigo.append(dato[2])
            cantidad.append(dato[3])
            remitente.append(dato[4])
            condicion.append(dato[5])
            niveleducativo.append(dato[6])
            añorecepcion.append(dato[7])
        df_datos = {
            'Titulo':titulo, 'Categoria': categoria, 'Codigo': codigo,
            'Cantidad': cantidad, 'Remitente': remitente,'Condicion': condicion,
            'Nivel Educativo': niveleducativo,'Año de Entrega': añorecepcion}
        
        fecha = str(strftime('%d-%m-%y_%H-%M-%S'))
        name_xlsx = f'LAMINAS {fecha}.xlsx'
        df = pd.DataFrame(df_datos)
        df.to_excel(name_xlsx, sheet_name='laminas')
        workbook = openpyxl.load_workbook(name_xlsx)
        sheet = workbook['laminas']
        sheet.delete_cols(1)
        workbook.save(name_xlsx)
        
    def save_bajalaminas(self):
        l_datos = self.bd.search_laminas('la.TipoId',24)
        i = -1
        titulo,categoria,codigo,cantidad = [],[],[],[]
        remitente,condicion,niveleducativo,añorecepcion= [],[],[],[]
        for dato in l_datos:
            titulo.append(dato[0])
            categoria.append(dato[1])
            codigo.append(dato[2])
            cantidad.append(dato[3])
            remitente.append(dato[4])
            condicion.append(dato[5])
            niveleducativo.append(dato[6])
            añorecepcion.append(dato[7])
        df_datos = {
            'Titulo':titulo, 'Categoria': categoria, 'Codigo': codigo,
            'Cantidad': cantidad, 'Remitente': remitente,'Condicion': condicion,
            'Nivel Educativo': niveleducativo,'Año de Entrega': añorecepcion}
        
        fecha = str(strftime('%d-%m-%y_%H-%M-%S'))
        name_xlsx = f'LAMINAS_BAJA {fecha}.xlsx'
        df = pd.DataFrame(df_datos)
        df.to_excel(name_xlsx, sheet_name='laminas_baja')
        workbook = openpyxl.load_workbook(name_xlsx)
        sheet = workbook['laminas_baja']
        sheet.delete_cols(1)
        workbook.save(name_xlsx)
        
    def save_nomina(self):
        l_alumnos = self.bd.show_alumnos()
        i = -1
        alumnoid,alumno,sexo = [],[],[]
        nivel,grado,seccion = [],[],[]
        for datos_alumno in l_alumnos:
            alumno.append(datos_alumno[0])
            sexo.append(datos_alumno[1])
            nivel.append(datos_alumno[2])
            grado.append(datos_alumno[3])
            seccion.append(datos_alumno[4])
            alumnoid.append(datos_alumno[5])
        dicionario_alumno = {
            'Alumno': alumno,'Sexo': sexo, 'Nivel': nivel,'Grado': grado, 
            'Seccion': seccion,'AlumnoId': alumnoid}
        fecha = str(strftime('%d-%m-%y_%H-%M-%S'))
        name_xlsx = f'ALUMNOS {fecha}.xlsx'
        data_alumnos = pd.DataFrame(dicionario_alumno)
        data_alumnos.to_excel(name_xlsx, sheet_name='alumnos')
        workbook = openpyxl.load_workbook(name_xlsx)
        sheet = workbook['alumnos']
        sheet.delete_cols(1)
        workbook.save(name_xlsx)
        
    def save_profesores(self):
        l_alumnos = self.bd.show_profesores()
        i = -1
        profesorid,profesor = [],[]
        celular,correo = [],[]
        for datos_alumno in l_alumnos:
            profesor.append(datos_alumno[0])
            correo.append(datos_alumno[1])
            celular.append(datos_alumno[2])
            profesorid.append(datos_alumno[3])
        dicionario_profesor = {
            'Profesor': profesor,'Correo': correo, 'Celular': celular,
            'Codigo': profesorid}
        fecha = str(strftime('%d-%m-%y_%H-%M-%S'))
        name_xlsx = f'PROFESORES {fecha}.xlsx'
        data_profesores = pd.DataFrame(dicionario_profesor)
        data_profesores.to_excel(name_xlsx, sheet_name='profesores')
        workbook = openpyxl.load_workbook(name_xlsx)
        sheet = workbook['profesores']
        sheet.delete_cols(1)
        workbook.save(name_xlsx)
        
    def save_pedidoslibros(self):
        pedidosalu = self.bd.showalu_pedidoslib()
        pedidospro = self.bd.showpro_pedidoslib()
        i = -1
        titulo,usuario,fechasalida,fechaentrada = [],[],[],[]
        cantidad,situacion,observacion,codigo,tipo= [],[],[],[],[]
        for datoalu in pedidosalu:
            titulo.append(datoalu[0])
            usuario.append(datoalu[1])
            fechasalida.append(datoalu[2])
            fechaentrada.append(datoalu[3])
            cantidad.append(datoalu[4])
            situacion.append(datoalu[5])
            observacion.append(datoalu[6])
            codigo.append(datoalu[7])
            tipo.append('Alumno')
            
        for datopro in pedidospro:
            titulo.append(datopro[0])
            usuario.append(datopro[1])
            fechasalida.append(datopro[2])
            fechaentrada.append(datopro[3])
            cantidad.append(datopro[4])
            situacion.append(datopro[5])
            observacion.append(datopro[6])
            codigo.append(datopro[7])
            tipo.append('Profesor')
        df_datos = {
            'Titulo': titulo,'Usuario': usuario,'Fecha de pedido': fechasalida,
            'Fecha de devolución': fechaentrada,'Situacion': situacion,'Cantidad': cantidad,
            'Observacion': observacion,'Codigo': codigo,'Tipo': tipo}
        
        fecha = str(strftime('%d-%m-%y_%H-%M-%S'))
        name_xlsx = f'LISTA_LIBROS_PEDIDOS {fecha}.xlsx'
        df = pd.DataFrame(df_datos)
        df.to_excel(name_xlsx, sheet_name='libros_pedidos')
        workbook = openpyxl.load_workbook(name_xlsx)
        sheet = workbook['libros_pedidos']
        sheet.delete_cols(1)
        workbook.save(name_xlsx)
        
    def save_pedidoslaminas(self):
        pedidosalu = self.bd.showalu_pedidoslam()
        pedidospro = self.bd.showpro_pedidoslam()
        i = -1
        titulo,usuario,fechasalida,fechaentrada = [],[],[],[]
        cantidad,situacion,observacion,codigo,tipo= [],[],[],[],[]
        for datoalu in pedidosalu:
            titulo.append(datoalu[0])
            usuario.append(datoalu[1])
            fechasalida.append(datoalu[2])
            fechaentrada.append(datoalu[3])
            cantidad.append(datoalu[4])
            situacion.append(datoalu[5])
            observacion.append(datoalu[6])
            codigo.append(datoalu[7])
            tipo.append('Alumno')
            
        for datopro in pedidospro:
            titulo.append(datopro[0])
            usuario.append(datopro[1])
            fechasalida.append(datopro[2])
            fechaentrada.append(datopro[3])
            cantidad.append(datopro[4])
            situacion.append(datopro[5])
            observacion.append(datopro[6])
            codigo.append(datopro[7])
            tipo.append('Profesor')
        df_datos = {
            'Titulo': titulo,'Usuario': usuario,'Fecha de pedido': fechasalida,
            'Fecha de devolución': fechaentrada,'Situacion': situacion,'Cantidad': cantidad,
            'Observacion': observacion,'Codigo': codigo,'Tipo': tipo}
        
        fecha = str(strftime('%d-%m-%y_%H-%M-%S'))
        name_xlsx = f'LISTA_LAMINAS_PEDIDOS {fecha}.xlsx'
        df = pd.DataFrame(df_datos)
        df.to_excel(name_xlsx, sheet_name='laminas_pedidos')
        workbook = openpyxl.load_workbook(name_xlsx)
        sheet = workbook['laminas_pedidos']
        sheet.delete_cols(1)
        workbook.save(name_xlsx)
        